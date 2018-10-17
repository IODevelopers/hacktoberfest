# -*- coding: utf-8 -*-
# Import `os` 
import os
# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
import argparse
from gooey import Gooey, GooeyParser
import message

@Gooey(dump_build_config=True, program_name="AutoValida",language='spanish')

def main():
    # Retrieve current working directory (`cwd`)
    cwd = os.getcwd()
    cwd

    # Change directory 
    # os.chdir("/home/rene/Proyectos/Validacion_Python")

    # List all files and directories in current directory
    print(os.listdir('.'))
    desc = "Aplicación que genera hojas de validación de nómina"
    file_help_msg = "Elige el archivo de la quincena a validar"

    my_cool_parser = GooeyParser(description=desc)
    my_cool_parser.add_argument("FileChooser", help=file_help_msg, widget="FileChooser")
    verbosity = my_cool_parser.add_mutually_exclusive_group()
    verbosity.add_argument('-t', '--verbozze', dest='verbose', action="store_true", help="Mostrar más detalles")
    verbosity.add_argument('-q', '--quiet', dest='quiet', action="store_true", help="Sólo muestra salida en error")

    args = my_cool_parser.parse_args()
    # display_message()

    def here_is_smore():
        pass
    # Load in the workbook
    wb = load_workbook(input("Archivo: "))

    # Get sheet names
    print(wb.sheetnames)

if __name__ == '__main__':
  main()
